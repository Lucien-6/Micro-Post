"""
Data management module.

This module handles data filtering, merging, and management
for the trajectory summary Excel file.
"""

import os
import re
import pandas as pd
import numpy as np
from typing import Dict, List, Set, Optional
import random
from openpyxl.utils import get_column_letter

from src.core.data_loader import DataLoader
from src.utils.logger import get_logger


class DataManager:
    """
    Manages trajectory data filtering, merging, and persistence.

    Attributes:
        data_loader: DataLoader instance for reading files.
        min_duration: Minimum tracking duration threshold (seconds).
        objects: Dictionary of object DataFrames keyed by new Object_ID.
        source_info: List of source information for each object.
        excluded_objects: Set of excluded object IDs.
        summary_path: Path to the summary Excel file.
    """

    def __init__(self, data_loader: DataLoader):
        """
        Initialize DataManager with a DataLoader instance.

        Args:
            data_loader: Configured DataLoader instance.
        """
        self.logger = get_logger()
        self.data_loader = data_loader
        self.min_duration: int = 10
        self.objects: Dict[str, pd.DataFrame] = {}
        self.source_info: List[Dict[str, str]] = []
        self.excluded_objects: Set[str] = set()
        self.summary_path: str = ""

        # Fitting info
        self.fitting_model: str = ""
        self.fitting_range: str = ""

    def set_fitting_info(self, model: str, fit_range: str) -> None:
        """
        Set fitting model and range information.

        Args:
            model: Fitting model name.
            fit_range: Fitting range string (e.g., "0 - 15").
        """
        self.fitting_model = model
        self.fitting_range = fit_range
        self.logger.info(
            f"Fitting info set: model={model}, range={fit_range} s"
        )

    def _auto_adjust_column_width(self, writer: pd.ExcelWriter):
        """
        Auto-adjust column widths for all sheets in the workbook.

        Args:
            writer: pandas ExcelWriter with openpyxl engine.
        """
        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except (TypeError, AttributeError):
                        pass

                # Add padding and set minimum/maximum width
                adjusted_width = min(max(max_length + 2, 8), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    def filter_and_merge(self, min_duration: int) -> Dict[str, int]:
        """
        Filter objects by duration and merge into unified dataset.

        Args:
            min_duration: Minimum tracking duration in seconds.

        Returns:
            Dictionary containing statistics:
            - total_files: Number of data files processed
            - total_objects: Total objects before filtering
            - passed_objects: Objects that passed filtering
        """
        self.min_duration = min_duration
        self.objects.clear()
        self.source_info.clear()
        self.excluded_objects.clear()

        object_counter = 1
        total_objects_before = 0
        total_files = len(self.data_loader.excel_files)

        for filepath in self.data_loader.excel_files:
            file_objects = self.data_loader.load_object_sheets(filepath)
            filename = os.path.basename(filepath)
            total_objects_before += len(file_objects)

            for original_id, df in file_objects.items():
                # Calculate tracking duration (number of rows - 1)
                duration = len(df) - 1 if len(df) > 0 else 0

                if duration >= min_duration:
                    new_id = f"Object_{object_counter}"
                    self.objects[new_id] = df.copy()
                    self.source_info.append({
                        "Object_ID": new_id,
                        "Source_File": filename,
                        "Original_Object": original_id
                    })
                    self.logger.debug(
                        f"Added {new_id}: {filename}/{original_id} "
                        f"(duration: {duration}s)"
                    )
                    object_counter += 1

        passed_count = len(self.objects)
        self.logger.info(
            f"Filtered and merged {passed_count} objects "
            f"(min_duration: {min_duration}s)"
        )

        return {
            "total_files": total_files,
            "total_objects": total_objects_before,
            "passed_objects": passed_count
        }

    def save_summary_excel(self) -> str:
        """
        Save merged data to Trajectories_Summary.xlsx.

        Creates an Excel file with:
        - Parameters sheet: Experiment parameters + Data Path
        - Source sheet: Object source information
        - Object_N sheets: Individual object trajectory data

        Returns:
            Path to the saved Excel file.
        """
        self.summary_path = os.path.join(
            self.data_loader.root_folder,
            "Trajectories_Summary.xlsx"
        )

        with pd.ExcelWriter(
            self.summary_path,
            engine="openpyxl"
        ) as writer:
            # 1. Parameters sheet
            params = self.data_loader.get_reference_parameters()
            params_data = [
                {"Parameter": "Data Path", "Value": self.data_loader.root_folder}
            ]
            for param_name, param_value in params.items():
                params_data.append({
                    "Parameter": param_name,
                    "Value": param_value
                })
            params_df = pd.DataFrame(params_data)
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

            # 2. Source sheet
            source_df = pd.DataFrame(self.source_info)
            source_df.to_excel(writer, sheet_name="Source", index=False)

            # 3. Summary sheet (placeholder, will be filled after analysis)
            summary_placeholder = pd.DataFrame({"Note": ["Analysis pending"]})
            summary_placeholder.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            # 4. Object sheets
            for obj_id in sorted(
                self.objects.keys(),
                key=lambda x: int(x.split("_")[1])
            ):
                df = self.objects[obj_id]
                df.to_excel(writer, sheet_name=obj_id, index=False)

            # Auto-adjust column widths
            self._auto_adjust_column_width(writer)

        self.logger.info(f"Summary saved to: {self.summary_path}")
        return self.summary_path

    def exclude_objects(self, object_ids: List[str]) -> int:
        """
        Exclude specified objects from the dataset.

        Removes objects from memory and updates the Excel file.

        Args:
            object_ids: List of object IDs to exclude (e.g., ["Object_3", "Object_7"]).

        Returns:
            Number of objects actually excluded.
        """
        excluded_count = 0

        for obj_id in object_ids:
            if obj_id in self.objects and obj_id not in self.excluded_objects:
                del self.objects[obj_id]
                self.excluded_objects.add(obj_id)
                excluded_count += 1
                self.logger.info(f"Excluded: {obj_id}")

        # Update source_info
        self.source_info = [
            info for info in self.source_info
            if info["Object_ID"] not in self.excluded_objects
        ]

        # Update Excel file
        if excluded_count > 0 and self.summary_path:
            self._update_excel_after_exclusion()

        return excluded_count

    def _update_excel_after_exclusion(self):
        """Update Excel file after excluding objects."""
        try:
            with pd.ExcelWriter(
                self.summary_path,
                engine="openpyxl"
            ) as writer:
                # Parameters sheet
                params = self.data_loader.get_reference_parameters()
                params_data = [
                    {"Parameter": "Data Path", "Value": self.data_loader.root_folder}
                ]
                for param_name, param_value in params.items():
                    params_data.append({
                        "Parameter": param_name,
                        "Value": param_value
                    })
                params_df = pd.DataFrame(params_data)
                params_df.to_excel(writer, sheet_name="Parameters", index=False)

                # Source sheet (updated)
                source_df = pd.DataFrame(self.source_info)
                source_df.to_excel(writer, sheet_name="Source", index=False)

                # Summary placeholder
                summary_placeholder = pd.DataFrame({"Note": ["Analysis pending"]})
                summary_placeholder.to_excel(
                    writer,
                    sheet_name="Summary",
                    index=False
                )

                # Object sheets (remaining)
                for obj_id in sorted(
                    self.objects.keys(),
                    key=lambda x: int(x.split("_")[1])
                ):
                    df = self.objects[obj_id]
                    df.to_excel(writer, sheet_name=obj_id, index=False)

                # Auto-adjust column widths
                self._auto_adjust_column_width(writer)

            self.logger.info("Excel file updated after exclusion")

        except Exception as e:
            self.logger.error(f"Failed to update Excel after exclusion: {e}")

    def parse_exclude_input(self, input_str: str) -> List[int]:
        """
        Parse user input for object exclusion.

        Supports formats:
        - Single number: "5"
        - Comma-separated: "3, 7, 12"
        - Range: "5-10"
        - Mixed: "3, 7-10, 15"

        Args:
            input_str: User input string.

        Returns:
            Sorted list of unique object numbers.
        """
        numbers = set()

        # Remove whitespace and split by comma
        parts = [p.strip() for p in input_str.split(",")]

        for part in parts:
            if not part:
                continue

            # Check for range pattern
            range_match = re.match(r"(\d+)\s*-\s*(\d+)", part)
            if range_match:
                start = int(range_match.group(1))
                end = int(range_match.group(2))
                numbers.update(range(start, end + 1))
            elif part.isdigit():
                numbers.add(int(part))

        return sorted(numbers)

    def get_object_count(self) -> int:
        """
        Get the current number of valid objects.

        Returns:
            Number of objects not excluded.
        """
        return len(self.objects)

    def get_object_data(self, object_id: str) -> Optional[pd.DataFrame]:
        """
        Get trajectory data for a specific object.

        Args:
            object_id: Object identifier (e.g., "Object_1").

        Returns:
            DataFrame with trajectory data, or None if not found.
        """
        return self.objects.get(object_id)

    def get_all_object_ids(self) -> List[str]:
        """
        Get list of all current object IDs.

        Returns:
            List of object IDs sorted numerically.
        """
        return sorted(
            self.objects.keys(),
            key=lambda x: int(x.split("_")[1])
        )

    def get_random_objects(self, n: int) -> List[str]:
        """
        Randomly select n objects from the dataset.

        Args:
            n: Number of objects to select.

        Returns:
            List of randomly selected object IDs.
        """
        available = list(self.objects.keys())
        n = min(n, len(available))
        return random.sample(available, n)

    def update_object_data(self, object_id: str, df: pd.DataFrame):
        """
        Update trajectory data for a specific object.

        Args:
            object_id: Object identifier.
            df: Updated DataFrame.
        """
        if object_id in self.objects:
            self.objects[object_id] = df

    def save_analysis_results(self, summary_df: pd.DataFrame):
        """
        Save analysis results to the summary Excel file.

        Updates individual object sheets and the Summary sheet.

        Args:
            summary_df: DataFrame containing summary statistics.
        """
        try:
            with pd.ExcelWriter(
                self.summary_path,
                engine="openpyxl"
            ) as writer:
                # Parameters sheet
                params = self.data_loader.get_reference_parameters()
                params_data = [
                    {"Parameter": "Data Path", "Value": self.data_loader.root_folder}
                ]
                for param_name, param_value in params.items():
                    params_data.append({
                        "Parameter": param_name,
                        "Value": param_value
                    })

                # Add fitting info to parameters
                if self.fitting_model:
                    params_data.append({
                        "Parameter": "Fitting Model",
                        "Value": self.fitting_model
                    })
                if self.fitting_range:
                    params_data.append({
                        "Parameter": "Fitting Range (s)",
                        "Value": self.fitting_range
                    })

                params_df = pd.DataFrame(params_data)
                params_df.to_excel(writer, sheet_name="Parameters", index=False)

                # Source sheet
                source_df = pd.DataFrame(self.source_info)
                source_df.to_excel(writer, sheet_name="Source", index=False)

                # Summary sheet (with analysis results)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

                # Object sheets (with analysis columns)
                for obj_id in sorted(
                    self.objects.keys(),
                    key=lambda x: int(x.split("_")[1])
                ):
                    df = self.objects[obj_id]
                    df.to_excel(writer, sheet_name=obj_id, index=False)

                # Auto-adjust column widths
                self._auto_adjust_column_width(writer)

            self.logger.info("Analysis results saved to Excel")

        except Exception as e:
            self.logger.error(f"Failed to save analysis results: {e}")
            raise
